#!/usr/bin/env python3
"""
SAFESCORING.IO - Comparaison Excel vs IA
Compare les scores du fichier Excel avec les scores calculés par l'IA
"""

import requests
from openpyxl import load_workbook
import re

# Configuration Supabase
SUPABASE_URL = 'https://ajdncttomdqojlozxjxu.supabase.co'
SUPABASE_KEY = 'REVOKED_ROTATE_ON_DASHBOARD'

HEADERS = {
    'apikey': SUPABASE_KEY,
    'Authorization': f'Bearer {SUPABASE_KEY}',
}

EXCEL_PATH = r'c:\Users\alexa\Desktop\SafeScoring\data\SAFE_SCORING_V7_FINAL.xlsx'


def load_excel_scores():
    """Charge les scores depuis le fichier Excel"""
    print("📊 Chargement du fichier Excel...")
    
    wb = load_workbook(EXCEL_PATH, data_only=True)
    
    # Chercher la feuille avec les scores
    scores = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"   Feuille: {sheet_name}")
        
        # Chercher les colonnes avec les scores
        header_row = None
        name_col = None
        score_col = None
        s_col = None
        a_col = None
        f_col = None
        e_col = None
        
        # Parcourir les premières lignes pour trouver les headers
        for row_idx in range(1, 10):
            for col_idx in range(1, 20):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    cell_str = str(cell_value).lower().strip()
                    
                    if 'product' in cell_str or 'produit' in cell_str or 'name' in cell_str or 'nom' in cell_str:
                        header_row = row_idx
                        name_col = col_idx
                    elif cell_str in ['safe', 'safe score', 'score', 'total']:
                        score_col = col_idx
                    elif cell_str == 's' or 'security' in cell_str:
                        s_col = col_idx
                    elif cell_str == 'a' or 'adversity' in cell_str:
                        a_col = col_idx
                    elif cell_str == 'f' or 'fidelity' in cell_str:
                        f_col = col_idx
                    elif cell_str == 'e' or 'efficiency' in cell_str:
                        e_col = col_idx
        
        if header_row and name_col:
            print(f"      Header trouvé ligne {header_row}, colonne nom: {name_col}")
            if score_col:
                print(f"      Colonne score: {score_col}")
            
            # Lire les données
            for row_idx in range(header_row + 1, ws.max_row + 1):
                name = ws.cell(row=row_idx, column=name_col).value
                if not name:
                    continue
                
                name = str(name).strip()
                
                product_scores = {}
                
                if score_col:
                    val = ws.cell(row=row_idx, column=score_col).value
                    if val is not None:
                        try:
                            # Convertir en pourcentage si nécessaire
                            if isinstance(val, (int, float)):
                                if val <= 1:
                                    product_scores['SAFE'] = val * 100
                                else:
                                    product_scores['SAFE'] = val
                        except:
                            pass
                
                if s_col:
                    val = ws.cell(row=row_idx, column=s_col).value
                    if val is not None:
                        try:
                            if isinstance(val, (int, float)):
                                product_scores['S'] = val * 100 if val <= 1 else val
                        except:
                            pass
                
                if a_col:
                    val = ws.cell(row=row_idx, column=a_col).value
                    if val is not None:
                        try:
                            if isinstance(val, (int, float)):
                                product_scores['A'] = val * 100 if val <= 1 else val
                        except:
                            pass
                
                if f_col:
                    val = ws.cell(row=row_idx, column=f_col).value
                    if val is not None:
                        try:
                            if isinstance(val, (int, float)):
                                product_scores['F'] = val * 100 if val <= 1 else val
                        except:
                            pass
                
                if e_col:
                    val = ws.cell(row=row_idx, column=e_col).value
                    if val is not None:
                        try:
                            if isinstance(val, (int, float)):
                                product_scores['E'] = val * 100 if val <= 1 else val
                        except:
                            pass
                
                if product_scores:
                    scores[name] = product_scores
    
    print(f"   ✅ {len(scores)} produits avec scores trouvés dans Excel")
    return scores


def load_ai_scores():
    """Charge les scores IA depuis Supabase"""
    print("\n🤖 Chargement des scores IA depuis Supabase...")
    
    r = requests.get(
        f'{SUPABASE_URL}/rest/v1/products?select=name,scores',
        headers=HEADERS
    )
    products = r.json()
    
    scores = {}
    for p in products:
        if p.get('scores') and p['scores'].get('full'):
            scores[p['name']] = p['scores']['full']
    
    print(f"   ✅ {len(scores)} produits avec scores IA")
    return scores


def normalize_name(name):
    """Normalise un nom de produit pour la comparaison"""
    name = name.lower().strip()
    # Supprimer les caractères spéciaux
    name = re.sub(r'[^\w\s]', '', name)
    # Supprimer les espaces multiples
    name = re.sub(r'\s+', ' ', name)
    return name


def find_match(excel_name, ai_names):
    """Trouve le meilleur match pour un nom Excel dans les noms IA"""
    excel_norm = normalize_name(excel_name)
    
    # Match exact
    for ai_name in ai_names:
        if normalize_name(ai_name) == excel_norm:
            return ai_name
    
    # Match partiel
    for ai_name in ai_names:
        ai_norm = normalize_name(ai_name)
        if excel_norm in ai_norm or ai_norm in excel_norm:
            return ai_name
    
    # Match par mots clés
    excel_words = set(excel_norm.split())
    best_match = None
    best_score = 0
    
    for ai_name in ai_names:
        ai_words = set(normalize_name(ai_name).split())
        common = len(excel_words & ai_words)
        if common > best_score and common >= 1:
            best_score = common
            best_match = ai_name
    
    return best_match


def compare_scores(excel_scores, ai_scores):
    """Compare les scores Excel et IA"""
    print("\n" + "=" * 80)
    print("📊 COMPARAISON EXCEL vs IA")
    print("=" * 80)
    
    comparisons = []
    matched = 0
    unmatched_excel = []
    
    ai_names = list(ai_scores.keys())
    
    for excel_name, excel_data in excel_scores.items():
        ai_name = find_match(excel_name, ai_names)
        
        if ai_name and ai_name in ai_scores:
            matched += 1
            ai_data = ai_scores[ai_name]
            
            comparison = {
                'excel_name': excel_name,
                'ai_name': ai_name,
                'excel_safe': excel_data.get('SAFE'),
                'ai_safe': ai_data.get('SAFE'),
                'diff': None
            }
            
            if comparison['excel_safe'] is not None and comparison['ai_safe'] is not None:
                comparison['diff'] = comparison['ai_safe'] - comparison['excel_safe']
            
            comparisons.append(comparison)
        else:
            unmatched_excel.append(excel_name)
    
    print(f"\n✅ {matched} produits matchés")
    print(f"❌ {len(unmatched_excel)} produits Excel non trouvés dans IA")
    
    # Afficher les comparaisons
    print("\n" + "-" * 80)
    print(f"{'Produit':<35} | {'Excel':>8} | {'IA':>8} | {'Diff':>8} | Status")
    print("-" * 80)
    
    # Trier par différence absolue
    comparisons.sort(key=lambda x: abs(x['diff']) if x['diff'] is not None else 0, reverse=True)
    
    total_diff = 0
    count_diff = 0
    
    for c in comparisons:
        excel_str = f"{c['excel_safe']:.1f}%" if c['excel_safe'] is not None else "N/A"
        ai_str = f"{c['ai_safe']:.1f}%" if c['ai_safe'] is not None else "N/A"
        
        if c['diff'] is not None:
            diff_str = f"{c['diff']:+.1f}%"
            total_diff += abs(c['diff'])
            count_diff += 1
            
            if abs(c['diff']) < 5:
                status = "✅ OK"
            elif abs(c['diff']) < 15:
                status = "⚠️ Écart"
            else:
                status = "❌ Grand écart"
        else:
            diff_str = "N/A"
            status = "⚪ N/A"
        
        print(f"{c['excel_name'][:35]:<35} | {excel_str:>8} | {ai_str:>8} | {diff_str:>8} | {status}")
    
    # Statistiques
    print("\n" + "=" * 80)
    print("📈 STATISTIQUES")
    print("=" * 80)
    
    if count_diff > 0:
        avg_diff = total_diff / count_diff
        print(f"   Écart moyen absolu: {avg_diff:.1f}%")
        
        # Compter les catégories
        ok_count = sum(1 for c in comparisons if c['diff'] is not None and abs(c['diff']) < 5)
        warn_count = sum(1 for c in comparisons if c['diff'] is not None and 5 <= abs(c['diff']) < 15)
        bad_count = sum(1 for c in comparisons if c['diff'] is not None and abs(c['diff']) >= 15)
        
        print(f"   ✅ OK (<5%): {ok_count} produits")
        print(f"   ⚠️ Écart (5-15%): {warn_count} produits")
        print(f"   ❌ Grand écart (>15%): {bad_count} produits")
    
    # Produits non matchés
    if unmatched_excel:
        print(f"\n❌ Produits Excel non trouvés ({len(unmatched_excel)}):")
        for name in unmatched_excel[:10]:
            print(f"   - {name}")
        if len(unmatched_excel) > 10:
            print(f"   ... et {len(unmatched_excel) - 10} autres")


def main():
    print("=" * 80)
    print("🔄 SAFESCORING - Comparaison Excel vs IA")
    print("=" * 80)
    
    excel_scores = load_excel_scores()
    ai_scores = load_ai_scores()
    
    if excel_scores and ai_scores:
        compare_scores(excel_scores, ai_scores)
    else:
        print("❌ Impossible de charger les scores")


if __name__ == '__main__':
    main()
